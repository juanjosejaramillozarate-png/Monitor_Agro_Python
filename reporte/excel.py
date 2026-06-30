"""Genera el libro Excel comercial con estructura lista para análisis."""

from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import COLORES_INTERFAZ


_VERDE = COLORES_INTERFAZ["acento"].replace("#", "")
_VERDE_SUAVE = COLORES_INTERFAZ["sidebar"].replace("#", "")
_TEXTO = COLORES_INTERFAZ["texto"].replace("#", "")
_BORDE = COLORES_INTERFAZ["borde"].replace("#", "")
_MORADO = "7C3AED"
_AZUL = "0E7490"


_COLUMNAS_ES = {
    "semana_cierre": "Cierre del periodo",
    "fecha_dato": "Fecha del dato",
    "variable": "Código de variable",
    "indicador": "Indicador",
    "valor": "Valor",
    "unidad": "Unidad",
    "cadencia": "Cadencia",
    "cambio_semanal_pct": "Cambio semanal (%)",
    "cambio_4_semanas_pct": "Cambio 4 semanas (%)",
    "cambio_mensual_pct": "Cambio mensual (%)",
    "cambio_interanual_pct": "Cambio interanual (%)",
    "indice_base_100": "Índice base 100",
    "fuente": "Fuente",
    "alcance_geografico": "Alcance geográfico",
}

_COLUMNAS_EN = {
    "semana_cierre": "Period close",
    "fecha_dato": "Data date",
    "variable": "Variable code",
    "indicador": "Indicator",
    "valor": "Value",
    "unidad": "Unit",
    "cadencia": "Frequency",
    "cambio_semanal_pct": "Weekly change (%)",
    "cambio_4_semanas_pct": "4-week change (%)",
    "cambio_mensual_pct": "Monthly change (%)",
    "cambio_interanual_pct": "Year-over-year change (%)",
    "indice_base_100": "Base-100 index",
    "fuente": "Source",
    "alcance_geografico": "Geographic scope",
}


def _estilo_cabecera(celda) -> None:
    celda.fill = PatternFill("solid", fgColor=_VERDE)
    celda.font = Font(color="FFFFFF", bold=True)
    celda.alignment = Alignment(vertical="center", wrap_text=True)


def _preparar_tabla(tabla: pd.DataFrame, idioma: str) -> pd.DataFrame:
    salida = tabla.copy()
    for columna in ("semana_cierre", "fecha_dato"):
        salida[columna] = pd.to_datetime(salida[columna]).dt.tz_localize(None)
    columnas = _COLUMNAS_EN if idioma == "en" else _COLUMNAS_ES
    return salida.rename(columns=columnas)


def _crear_resumen(writer, tabla: pd.DataFrame, idioma: str) -> None:
    wb = writer.book
    nombre = "Summary" if idioma == "en" else "Resumen"
    ws = wb.create_sheet(nombre, 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A10"

    titulo = "Monitor Agro Colombia — Commercial data" if idioma == "en" else "Monitor Agro Colombia — Datos comerciales"
    subtitulo = (
        "Reusable evidence with date, unit, source and frequency."
        if idioma == "en"
        else "Evidencia reutilizable con fecha, unidad, fuente y cadencia."
    )
    ws.merge_cells("A1:F1")
    ws["A1"] = titulo
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=_VERDE)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:F2")
    ws["A2"] = subtitulo
    ws["A2"].font = Font(size=10, color=_TEXTO)

    etiquetas = (
        ["Data range", "Rows", "Indicators", "Generated"]
        if idioma == "en"
        else ["Rango de datos", "Filas", "Indicadores", "Generado"]
    )
    fechas = pd.to_datetime(tabla["fecha_dato"])
    valores = [
        f"{fechas.min():%d/%m/%Y} – {fechas.max():%d/%m/%Y}",
        len(tabla),
        tabla["variable"].nunique(),
        f"{date.today():%d/%m/%Y}",
    ]
    for columna, (etiqueta, valor) in enumerate(zip(etiquetas, valores), start=1):
        celda_etiqueta = ws.cell(4, columna, etiqueta)
        celda_valor = ws.cell(5, columna, valor)
        celda_etiqueta.fill = PatternFill("solid", fgColor=_VERDE_SUAVE)
        celda_etiqueta.font = Font(bold=True, color=_VERDE)
        celda_valor.font = Font(size=12, bold=True, color=_TEXTO)
        for celda in (celda_etiqueta, celda_valor):
            celda.alignment = Alignment(vertical="center", wrap_text=True)
            celda.border = Border(bottom=Side(style="thin", color=_BORDE))
    ws.row_dimensions[5].height = 30

    ws.merge_cells("A7:F7")
    ws["A7"] = (
        "Use ‘Commercial series’ to filter and analyse observations. Blank changes mean that the cadence does not support that comparison."
        if idioma == "en"
        else "Use ‘Series comerciales’ para filtrar y analizar observaciones. Los cambios vacíos indican que la cadencia no permite esa comparación."
    )
    ws["A7"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A7"].font = Font(size=9, italic=True, color=_TEXTO)
    ws.row_dimensions[7].height = 32

    encabezados = (
        ["Variable code", "Indicator", "Latest data", "Value", "Unit", "Source"]
        if idioma == "en"
        else ["Código", "Indicador", "Último dato", "Valor", "Unidad", "Fuente"]
    )
    for col, valor in enumerate(encabezados, start=1):
        ws.cell(9, col, valor)
        _estilo_cabecera(ws.cell(9, col))

    ultimos = (
        tabla.sort_values("fecha_dato")
        .groupby("variable", as_index=False, sort=False)
        .tail(1)
        .sort_values("variable")
    )
    for fila_excel, (_, fila) in enumerate(ultimos.iterrows(), start=10):
        valores_fila = [
            fila["variable"],
            fila["indicador"],
            pd.Timestamp(fila["fecha_dato"]).to_pydatetime(),
            float(fila["valor"]),
            fila["unidad"],
            fila["fuente"],
        ]
        for col, valor in enumerate(valores_fila, start=1):
            ws.cell(fila_excel, col, valor)
        ws.cell(fila_excel, 3).number_format = "dd/mm/yyyy"
        ws.cell(fila_excel, 4).number_format = "#,##0.0"
        for celda in ws[fila_excel][:6]:
            celda.alignment = Alignment(vertical="top", wrap_text=True)
            celda.border = Border(bottom=Side(style="thin", color=_BORDE))

    mensuales = tabla[tabla["variable"].isin(["produccion_nacional", "exportaciones_cafe"])].copy()
    if not mensuales.empty:
        mensuales["Mes"] = pd.to_datetime(mensuales["fecha_dato"]).dt.strftime("%Y-%m")
        pivote = mensuales.pivot_table(index="Mes", columns="variable", values="valor", aggfunc="last").tail(12)
        ws["H9"] = "Month" if idioma == "en" else "Mes"
        ws["I9"] = "Production" if idioma == "en" else "Producción"
        ws["J9"] = "Exports" if idioma == "en" else "Exportaciones"
        for celda in ws[9][7:10]:
            _estilo_cabecera(celda)
        for fila_excel, (mes, fila) in enumerate(pivote.iterrows(), start=10):
            ws.cell(fila_excel, 8, mes)
            ws.cell(fila_excel, 9, fila.get("produccion_nacional"))
            ws.cell(fila_excel, 10, fila.get("exportaciones_cafe"))
            ws.cell(fila_excel, 9).number_format = "#,##0.0"
            ws.cell(fila_excel, 10).number_format = "#,##0.0"
        fin = 9 + len(pivote)
        grafico = BarChart()
        grafico.type = "col"
        grafico.style = 10
        grafico.title = "Production and exports — last 12 months" if idioma == "en" else "Producción y exportaciones — últimos 12 meses"
        grafico.y_axis.title = "Thousand 60-kg bags" if idioma == "en" else "Miles de sacos de 60 kg"
        grafico.add_data(Reference(ws, min_col=9, max_col=10, min_row=9, max_row=fin), titles_from_data=True)
        grafico.set_categories(Reference(ws, min_col=8, min_row=10, max_row=fin))
        if len(grafico.series) >= 2:
            grafico.series[0].tx = SeriesLabel(v="Production" if idioma == "en" else "Producción")
            grafico.series[1].tx = SeriesLabel(v="Exports" if idioma == "en" else "Exportaciones")
            grafico.series[0].graphicalProperties.solidFill = _MORADO
            grafico.series[1].graphicalProperties.solidFill = _AZUL
        grafico.gapWidth = 60
        grafico.x_axis.tickLblSkip = 2
        grafico.height = 8
        grafico.width = 15
        ws.add_chart(grafico, "L4")

    anchos = {"A": 29, "B": 34, "C": 16, "D": 16, "E": 26, "F": 20, "H": 12, "I": 15, "J": 15}
    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho


def _crear_diccionario(writer, idioma: str) -> None:
    wb = writer.book
    nombre = "Data dictionary" if idioma == "en" else "Diccionario"
    ws = wb.create_sheet(nombre)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    definiciones_es = {
        "Cierre del periodo": "Cierre semanal o fecha mensual usada para ordenar la observación.",
        "Fecha del dato": "Fecha real publicada por la fuente.",
        "Código de variable": "Identificador estable de la serie dentro del proyecto.",
        "Indicador": "Nombre legible de la serie.",
        "Valor": "Observación numérica en la unidad indicada.",
        "Unidad": "Unidad original o legible de la observación.",
        "Cadencia": "Frecuencia y tratamiento temporal de la serie.",
        "Cambio semanal (%)": "Cambio frente al cierre semanal anterior.",
        "Cambio 4 semanas (%)": "Cambio frente al cierre disponible cuatro semanas atrás.",
        "Cambio mensual (%)": "Cambio frente al mes publicado anterior.",
        "Cambio interanual (%)": "Cambio frente al mismo mes o periodo del año anterior.",
        "Índice base 100": "Valor relativo respecto a la primera observación disponible desde 2023.",
        "Fuente": "Proveedor o institución de origen.",
        "Alcance geográfico": "Nivel territorial al que corresponde el dato.",
    }
    definiciones_en = {
        "Period close": "Weekly close or monthly date used to order the observation.",
        "Data date": "Actual date published by the source.",
        "Variable code": "Stable identifier used for the series within the project.",
        "Indicator": "Human-readable series name.",
        "Value": "Numeric observation in the stated unit.",
        "Unit": "Original or human-readable unit of the observation.",
        "Frequency": "Publication frequency and time treatment of the series.",
        "Weekly change (%)": "Change from the previous weekly close.",
        "4-week change (%)": "Change from the available close four weeks earlier.",
        "Monthly change (%)": "Change from the previously published month.",
        "Year-over-year change (%)": "Change from the same month or period one year earlier.",
        "Base-100 index": "Relative value against the first available observation since 2023.",
        "Source": "Originating provider or institution.",
        "Geographic scope": "Territorial level represented by the observation.",
    }
    definiciones = definiciones_en if idioma == "en" else definiciones_es
    ws.append(["Field", "Description"] if idioma == "en" else ["Campo", "Descripción"])
    for celda in ws[1]:
        _estilo_cabecera(celda)
    for campo, descripcion in definiciones.items():
        ws.append([campo, descripcion])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 82
    for fila in ws.iter_rows(min_row=2):
        for celda in fila:
            celda.alignment = Alignment(vertical="top", wrap_text=True)


def generar_excel_comercial(tabla: pd.DataFrame, idioma: str = "es") -> bytes:
    """Devuelve un libro con resumen, tabla filtrable y diccionario."""
    if tabla.empty:
        raise ValueError("No hay datos comerciales para exportar")
    hoja_series = "Commercial series" if idioma == "en" else "Series comerciales"
    datos = _preparar_tabla(tabla, idioma)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        datos.to_excel(writer, index=False, sheet_name=hoja_series)
        ws = writer.book[hoja_series]
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        tabla_excel = Table(displayName="CommercialSeries", ref=ws.dimensions)
        tabla_excel.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tabla_excel)
        ws.row_dimensions[1].height = 32
        for celda in ws[1]:
            _estilo_cabecera(celda)
        for columna in (1, 2):
            for celda in ws.iter_cols(min_col=columna, max_col=columna, min_row=2):
                for item in celda:
                    item.number_format = "dd/mm/yyyy"
        for columna in range(8, 12):
            for celda in ws.iter_cols(min_col=columna, max_col=columna, min_row=2):
                for item in celda:
                    item.number_format = '0.0"%";[Red]-0.0"%"'
        for celda in ws.iter_cols(min_col=5, max_col=5, min_row=2):
            for item in celda:
                item.number_format = "#,##0.0"
        for celda in ws.iter_cols(min_col=12, max_col=12, min_row=2):
            for item in celda:
                item.number_format = "0.0"
        anchos = [16, 16, 24, 38, 16, 25, 34, 20, 22, 20, 23, 17, 18, 22]
        for indice, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[ws.cell(1, indice).column_letter].width = ancho
        _crear_resumen(writer, tabla, idioma)
        _crear_diccionario(writer, idioma)
        writer.book.active = 0
    return buffer.getvalue()
