from io import BytesIO
import unittest

import pandas as pd
from openpyxl import load_workbook

from reporte.excel import generar_excel_comercial


class ExcelComercialTests(unittest.TestCase):
    def _datos(self) -> pd.DataFrame:
        filas = []
        for indice, fecha in enumerate(pd.date_range("2026-01-01", periods=4, freq="MS")):
            for variable, indicador in [
                ("produccion_nacional", "Producción nacional de café"),
                ("exportaciones_cafe", "Exportaciones colombianas de café"),
            ]:
                filas.append(
                    {
                        "semana_cierre": fecha,
                        "fecha_dato": fecha,
                        "variable": variable,
                        "indicador": indicador,
                        "valor": 900.0 + indice,
                        "unidad": "miles de sacos de 60 kg",
                        "cadencia": "Mensual",
                        "cambio_semanal_pct": pd.NA,
                        "cambio_4_semanas_pct": pd.NA,
                        "cambio_mensual_pct": 1.2,
                        "cambio_interanual_pct": 4.5,
                        "indice_base_100": 101.0,
                        "fuente": "FNC",
                        "alcance_geografico": "COLOMBIA",
                    }
                )
        return pd.DataFrame(filas)

    def test_libro_incluye_resumen_tabla_y_diccionario(self) -> None:
        contenido = generar_excel_comercial(self._datos())

        libro = load_workbook(BytesIO(contenido))
        self.assertEqual(libro.sheetnames, ["Resumen", "Series comerciales", "Diccionario"])
        series = libro["Series comerciales"]
        self.assertEqual(series.freeze_panes, "A2")
        self.assertIn("CommercialSeries", series.tables)
        self.assertIsNone(series.auto_filter.ref)
        self.assertEqual(series.tables["CommercialSeries"].autoFilter.ref, "A1:N9")
        self.assertEqual(series["A2"].number_format, "dd/mm/yyyy")
        self.assertIn('%', series["J2"].number_format)
        self.assertFalse(series.sheet_view.showGridLines)

        resumen = libro["Resumen"]
        self.assertGreaterEqual(len(resumen._charts), 1)
        self.assertEqual(resumen["A1"].value, "Monitor Agro Colombia — Datos comerciales")

    def test_libro_rechaza_tabla_vacia(self) -> None:
        with self.assertRaisesRegex(ValueError, "No hay datos"):
            generar_excel_comercial(pd.DataFrame())


if __name__ == "__main__":
    unittest.main()
